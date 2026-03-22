#!/usr/bin/env python3

from __future__ import annotations

import json
import pathlib
import re
import sys
from collections import defaultdict
from dataclasses import dataclass

from openpyxl import load_workbook


NUMBER_WORDS = {
    'zero',
    'one',
    'two',
    'three',
    'four',
    'five',
    'six',
    'seven',
    'eight',
    'nine',
    'ten',
    'eleven',
    'twelve',
    'thirteen',
    'fourteen',
    'fifteen',
    'sixteen',
    'seventeen',
    'eighteen',
    'nineteen',
    'twenty',
    'thirty',
    'forty',
    'fifty',
    'sixty',
    'seventy',
    'eighty',
    'ninety',
    'hundred',
    'thousand',
    'million',
}

COLORS = {
    'red',
    'blue',
    'yellow',
    'green',
    'orange',
    'purple',
    'pink',
    'brown',
    'black',
    'white',
    'gray',
    'grey',
    'gold',
    'silver',
}

PLANET_DISTRACTORS = {
    'Mercury',
    'Venus',
    'Earth',
    'Mars',
    'Jupiter',
    'Saturn',
    'Uranus',
    'Neptune',
}

CHEMICAL_SYMBOL_DISTRACTORS = {
    'H2O',
    'CO2',
    'O',
    'Au',
    'Fe',
    'NaCl',
}


@dataclass(frozen=True)
class Entry:
    id: str
    category: str
    difficulty: str
    question: str
    correct_answer: str
    domains: frozenset[str]


@dataclass(frozen=True)
class ImportedQuestion:
    category: str
    difficulty: str
    question: str
    correct_answer: str
    direct_choices: tuple[str, str, str, str] | None
    domains: frozenset[str]


def norm(text: str) -> str:
    return re.sub(r'\s+', ' ', str(text).strip().lower())


def stringify_cell(value: object) -> str:
    if value is None:
        return ''
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f'{value:g}'
    return str(value)


def clean(value: object) -> str:
    return re.sub(r'\s+', ' ', stringify_cell(value).strip())


def normalize_difficulty(value: str) -> str:
    return clean(value).lower()


def answer_class(value: str) -> str:
    cleaned = clean(value)
    lower = cleaned.lower()
    words = re.findall(r"[A-Za-z0-9'.+-]+", cleaned)

    if re.fullmatch(r'\d{4}(?:\.0)?', lower):
        return 'year'
    if lower.endswith('.0') and lower[:-2].isdigit():
        return 'number'
    if lower in COLORS:
        return 'color'
    if lower in NUMBER_WORDS or re.fullmatch(r'[\d,.]+', lower):
        return 'number'
    if re.fullmatch(r'[A-Z]{1,4}\d{0,2}', cleaned):
        return 'symbol'
    if cleaned.isupper() and len(cleaned) <= 10:
        return 'acronym'
    if len(words) >= 2 and all((word[:1].isupper() or word.endswith('.')) for word in words if word):
        return 'person_or_proper'
    if len(words) >= 4:
        return 'phrase'
    if len(words) == 1:
        return 'single'
    return 'short'


def pattern(question: str) -> str:
    lowered = norm(question)
    rules = [
        ('capital of', 'capital'),
        ('chemical symbol', 'chemical-symbol'),
        ('who wrote', 'who-wrote'),
        ('who painted', 'who-painted'),
        ('in what year', 'year'),
        ('what year', 'year'),
        ('how many', 'how-many'),
        ('what color', 'color'),
        ('what instrument', 'instrument'),
        ('what does', 'what-does'),
        ('what is the name of', 'name-of'),
        ('who was', 'who-was'),
    ]
    for needle, label in rules:
        if needle in lowered:
            return label

    words = lowered.replace('?', '').split()
    return ' '.join(words[:3])


def keyword_domains(question: str) -> set[str]:
    lowered = norm(question)
    domains: set[str] = set()
    if any(token in lowered for token in ['planet', 'solar system', 'moon', 'star', 'galaxy', 'orbit', 'astronaut']):
        domains.add('astronomy')
    if any(token in lowered for token in ['chemical', 'acid', 'atom', 'molecule', 'gas', 'element', 'symbol']):
        domains.add('chemistry')
    if any(token in lowered for token in ['body', 'organ', 'cell', 'blood', 'bone', 'muscle']):
        domains.add('biology')
    if any(token in lowered for token in ['river', 'country', 'capital', 'mountain', 'ocean', 'sea', 'continent', 'desert']):
        domains.add('geography')
    if any(token in lowered for token in ['year', 'president', 'war', 'empire', 'wall', 'ancient', 'king', 'queen', 'revolution']):
        domains.add('history')
    if any(token in lowered for token in ['wrote', 'novel', 'poem', 'play', 'author', 'book']):
        domains.add('literature')
    if any(token in lowered for token in ['painted', 'artist', 'sculpture', 'museum', 'color']):
        domains.add('art')
    if any(token in lowered for token in ['instrument', 'song', 'singer', 'composer', 'band']):
        domains.add('music')
    return domains


def new_domains(category: str, question: str) -> frozenset[str]:
    mapping = {
        'geography': {'geography'},
        'science': {'science'},
        'history': {'history'},
        'literature': {'literature'},
        'art': {'art'},
        'music': {'music'},
        'movies': {'movies'},
        'sports': {'sports'},
        'technology': {'technology'},
        'food': {'food'},
        'nature': {'nature'},
        'general knowledge': {'general knowledge'},
    }
    domains = set(mapping.get(category.lower(), {category.lower()}))
    if 'science' in domains:
        domains.update({'astronomy', 'chemistry', 'biology', 'nature'})
    if 'nature' in domains:
        domains.update({'science', 'geography'})
    domains.update(keyword_domains(question))
    return frozenset(domains)


def legacy_domains(category: str, question: str) -> frozenset[str]:
    lowered = category.lower()
    domains: set[str] = set()
    if 'astronomy' in lowered:
        domains.update({'astronomy', 'science'})
    if 'chemistry' in lowered or 'physics' in lowered:
        domains.update({'chemistry', 'science'})
    if 'biology' in lowered:
        domains.update({'biology', 'science', 'nature'})
    if 'nature' in lowered or 'earth' in lowered:
        domains.update({'nature', 'science', 'geography'})
    if 'geography' in lowered:
        domains.add('geography')
    if 'history' in lowered or 'government' in lowered:
        domains.add('history')
    if 'literature' in lowered or 'language' in lowered:
        domains.add('literature')
    if 'art' in lowered:
        domains.add('art')
    if 'music' in lowered:
        domains.add('music')
    if 'math' in lowered or 'logic' in lowered:
        domains.update({'technology', 'general knowledge'})
    domains.update(keyword_domains(question))
    return frozenset(domains or {'general knowledge'})


def score_candidate(item: Entry, candidate: Entry) -> float:
    score = 0.0
    if answer_class(item.correct_answer) == answer_class(candidate.correct_answer):
        score += 8
    if pattern(item.question) == pattern(candidate.question):
        score += 6
    if item.domains & candidate.domains:
        score += 4
    score -= abs(len(item.correct_answer) - len(candidate.correct_answer)) / 8
    score -= abs(len(item.correct_answer.split()) - len(candidate.correct_answer.split())) * 1.5
    return score


def supplemental_entries(item: Entry) -> list[Entry]:
    lowered = norm(item.question)
    extras: list[Entry] = []

    if 'planet' in lowered or 'solar system' in lowered:
        extras.extend(
            Entry(
                id=f'supplemental-planet-{index}',
                category='Supplemental',
                difficulty=item.difficulty,
                question='Supplemental planet distractor',
                correct_answer=answer,
                domains=frozenset({'astronomy', 'science'}),
            )
            for index, answer in enumerate(sorted(PLANET_DISTRACTORS))
            if norm(answer) != norm(item.correct_answer)
        )

    if 'chemical symbol' in lowered:
        extras.extend(
            Entry(
                id=f'supplemental-symbol-{index}',
                category='Supplemental',
                difficulty=item.difficulty,
                question='Supplemental chemistry distractor',
                correct_answer=answer,
                domains=frozenset({'chemistry', 'science'}),
            )
            for index, answer in enumerate(sorted(CHEMICAL_SYMBOL_DISTRACTORS))
            if norm(answer) != norm(item.correct_answer)
        )

    return extras


def build_distractors(item: Entry, entries: list[Entry], existing_entries: list[Entry]) -> list[str]:
    by_category_diff: defaultdict[tuple[str, str], list[Entry]] = defaultdict(list)
    by_category: defaultdict[str, list[Entry]] = defaultdict(list)
    by_pattern: defaultdict[str, list[Entry]] = defaultdict(list)
    by_diff: defaultdict[str, list[Entry]] = defaultdict(list)
    by_class: defaultdict[str, list[Entry]] = defaultdict(list)
    existing_by_domain: defaultdict[str, list[Entry]] = defaultdict(list)
    existing_by_class: defaultdict[str, list[Entry]] = defaultdict(list)

    for candidate in entries:
        by_category_diff[(candidate.category, candidate.difficulty)].append(candidate)
        by_category[candidate.category].append(candidate)
        by_pattern[pattern(candidate.question)].append(candidate)
        by_diff[candidate.difficulty].append(candidate)
        by_class[answer_class(candidate.correct_answer)].append(candidate)

    for candidate in existing_entries:
        for domain in candidate.domains:
            existing_by_domain[domain].append(candidate)
        existing_by_class[answer_class(candidate.correct_answer)].append(candidate)

    class_name = answer_class(item.correct_answer)
    domain_existing = [candidate for domain in item.domains for candidate in existing_by_domain[domain]]

    pools: list[list[Entry]] = [
        supplemental_entries(item),
        [candidate for candidate in by_category_diff[(item.category, item.difficulty)] if answer_class(candidate.correct_answer) == class_name],
        [candidate for candidate in domain_existing if answer_class(candidate.correct_answer) == class_name],
        [candidate for candidate in existing_by_class[class_name] if pattern(item.question) == pattern(candidate.question)],
        [candidate for candidate in by_pattern[pattern(item.question)] if answer_class(candidate.correct_answer) == class_name],
        [candidate for candidate in by_category[item.category] if answer_class(candidate.correct_answer) == class_name],
        [candidate for candidate in by_diff[item.difficulty] if answer_class(candidate.correct_answer) == class_name],
        existing_by_class[class_name],
        by_category_diff[(item.category, item.difficulty)],
        domain_existing,
        by_category[item.category],
        by_class[class_name],
        by_diff[item.difficulty],
        entries,
    ]

    selected: list[str] = []
    seen_answers = {norm(item.correct_answer)}

    for pool in pools:
        unique_candidates = {}
        for candidate in pool:
            if candidate.id == item.id or norm(candidate.correct_answer) in seen_answers:
                continue
            unique_candidates[(candidate.id, norm(candidate.correct_answer))] = candidate

        ranked = sorted(
            unique_candidates.values(),
            key=lambda candidate: (-score_candidate(item, candidate), candidate.correct_answer.lower()),
        )

        for candidate in ranked:
            answer = clean(candidate.correct_answer)
            if norm(answer) in seen_answers:
                continue
            seen_answers.add(norm(answer))
            selected.append(answer)
            if len(selected) == 3:
                return selected

    return selected


def load_existing_entries(rows: list[dict]) -> list[Entry]:
    return [
        Entry(
            id=f'existing-{row["id"]}',
            category=clean(row['category']),
            difficulty=normalize_difficulty(row['difficulty']),
            question=clean(row['question']),
            correct_answer=clean(row['correctAnswer']),
            domains=legacy_domains(row['category'], row['question']),
        )
        for row in rows
    ]


def load_imported_questions(workbook_path: pathlib.Path) -> list[ImportedQuestion]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    entries: list[ImportedQuestion] = []
    seen_questions: set[str] = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        _, category, difficulty, question, answer, *rest = row
        if not question or not answer:
            continue

        normalized_question = norm(question)
        if normalized_question in seen_questions:
            continue
        seen_questions.add(normalized_question)

        direct_choices: tuple[str, str, str, str] | None = None
        cleaned_rest = [clean(value) for value in rest[:3]]
        if all(cleaned_rest):
            direct_choices = (clean(answer), cleaned_rest[0], cleaned_rest[1], cleaned_rest[2])
            if len({norm(choice) for choice in direct_choices}) != 4:
                raise RuntimeError(f'Duplicate authored choices for: {clean(question)}')

        entries.append(
            ImportedQuestion(
                category=clean(category),
                difficulty=normalize_difficulty(difficulty),
                question=clean(question),
                correct_answer=clean(answer),
                direct_choices=direct_choices,
                domains=new_domains(clean(category), clean(question)),
            )
        )

    return entries


def merge_bank(workbook_path: pathlib.Path, target_json: pathlib.Path) -> tuple[int, int, int, int]:
    existing_rows = json.loads(target_json.read_text())
    existing_entries = load_existing_entries(existing_rows)
    imported_questions = load_imported_questions(workbook_path)
    existing_by_question = {norm(row['question']): row for row in existing_rows}
    new_entries = [
        Entry(
            id=f'new-{index}',
            category=item.category,
            difficulty=item.difficulty,
            question=item.question,
            correct_answer=item.correct_answer,
            domains=item.domains,
        )
        for index, item in enumerate(imported_questions, start=1)
        if norm(item.question) not in existing_by_question
    ]

    next_numeric_id = max(row['id'] for row in existing_rows) + 1
    merged_rows = list(existing_rows)
    updated_count = 0
    added_count = 0

    for item in imported_questions:
        existing_row = existing_by_question.get(norm(item.question))
        if not existing_row or not item.direct_choices:
            continue

        existing_row['category'] = item.category
        existing_row['difficulty'] = item.difficulty
        existing_row['correctAnswer'] = item.correct_answer
        existing_row['choices'] = list(item.direct_choices)
        updated_count += 1

    imported_new_questions = [item for item in imported_questions if norm(item.question) not in existing_by_question]

    if len(imported_new_questions) != len(new_entries):
        raise RuntimeError('Imported question alignment failed while merging the bank.')

    for imported_question, entry in zip(imported_new_questions, new_entries):
        distractors = (
            list(imported_question.direct_choices[1:])
            if imported_question.direct_choices
            else build_distractors(entry, new_entries, existing_entries)
        )
        if len(distractors) < 3:
            raise RuntimeError(f'Could not build three distractors for: {entry.question}')

        merged_rows.append(
            {
                'id': next_numeric_id,
                'category': entry.category,
                'difficulty': entry.difficulty,
                'question': entry.question,
                'choices': [entry.correct_answer, *distractors],
                'correctAnswer': entry.correct_answer,
                'explanation': None,
            }
        )
        next_numeric_id += 1
        added_count += 1

    target_json.write_text(json.dumps(merged_rows, indent=2) + '\n')
    return len(existing_rows), len(merged_rows), updated_count, added_count


def main() -> int:
    if len(sys.argv) != 3:
        print('Usage: merge_question_bank.py <source.xlsx> <target.json>')
        return 1

    workbook_path = pathlib.Path(sys.argv[1]).expanduser()
    target_json = pathlib.Path(sys.argv[2]).expanduser()

    before_count, after_count, updated_count, added_count = merge_bank(workbook_path, target_json)
    print(
        f'Merged question bank: {before_count} -> {after_count} '
        f'(updated {updated_count}, added {added_count})'
    )
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
